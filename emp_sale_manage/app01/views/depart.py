from django.shortcuts import render,HttpResponse,redirect
from django.utils.safestring import mark_safe  # 标记安全字符串
from app01.utills.pagination import Pagination
from openpyxl import load_workbook
from app01 import models
# Create your views here.

# 部门管理
def depart_list(request):
    """ 部门列表 """
    queryset = models.Department.objects.all()

    context = {
        'queryset':queryset,
    }

    return render(request,'depart_list.html',context)


def depart_add(request):
    """ 添加部门 """
    if request.method == 'GET':
        return render(request,"depart_add.html")
    
    else:
        title = request.POST.get("title")
        models.Department.objects.create(title=title)

        # 重定向
        return redirect('http://127.0.0.1:8000/depart/list/')

def depart_delete(request):
    """ 删除部门 """
    nid = request.GET.get("nid")
    models.Department.objects.filter(id=nid).delete()
    return redirect('http://127.0.0.1:8000/depart/list/')

def depart_edit(request,nid):
    """ 编辑部门 """
    if request.method == 'GET':
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request,"depart_edit.html",{'row_object':row_object})

    else:
        new_title = request.POST.get("title")
        models.Department.objects.filter(id=nid).update(title=new_title)
        
        # 重定向
        return redirect('http://127.0.0.1:8000/depart/list/')

def depart_multi(request):
    """ 数据上传 """
    # 获取到文件对象
    file_object = request.FILES.get('exc')

    # 将对象传递给openpyxl
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 读取excel表中的数据
    for row in sheet.iter_rows(min_row=2):
        for col in row:
            text = col.value
            exist = models.Department.objects.filter(title=text).exists()
            if not exist:
                models.Department.objects.create(title=text)
   
    return redirect('http://127.0.0.1:8000/depart/list/')
